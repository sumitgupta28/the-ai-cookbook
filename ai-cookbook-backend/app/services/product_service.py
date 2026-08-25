from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import Settings
from app.models.product import Product
from app.repositories.product_repository import ProductRepository
from app.services.embedding_service import EmbeddingService
from app.vector.vector_store_adapter import VectorStoreAdapter


@dataclass
class ProductUploadResult:
    """Summary of imported, skipped, and invalid spreadsheet rows."""

    imported: int
    skipped: int
    errors: list[str]


class ProductService:
    """Import catalog spreadsheets and search the dedicated product vector store."""

    REQUIRED_HEADERS = [
        "ProductID",
        "Name",
        "Category",
        "Brand",
        "Description",
        "Price",
        "ImageUrl",
        "Rating",
        "StockCount",
    ]

    def __init__(self, session: Session, settings: Settings) -> None:
        """Initialize product, vector, and embedding dependencies."""
        self.session = session
        self.repository = ProductRepository(session)
        self.vector_store = VectorStoreAdapter(session)
        self.embeddings = EmbeddingService(settings)

    def ingest_products_xlsx(self, payload: bytes) -> ProductUploadResult:
        """Validate, upsert, embed, and report rows from an XLSX catalog."""
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            return ProductUploadResult(imported=0, skipped=0, errors=["Empty workbook"])

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ProductUploadResult(imported=0, skipped=0, errors=["Empty workbook"])

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        header_map = {header: index for index, header in enumerate(headers)}

        missing = [header for header in self.REQUIRED_HEADERS if header not in header_map]
        if missing:
            return ProductUploadResult(
                imported=0,
                skipped=0,
                errors=[f"Missing required columns: {', '.join(missing)}"],
            )

        imported = 0
        skipped = 0
        errors: list[str] = []

        for row_number, row in enumerate(rows[1:], start=2):
            try:
                product_id = str(row[header_map["ProductID"]] or "").strip()
                name = str(row[header_map["Name"]] or "").strip()
                if not product_id or not name:
                    skipped += 1
                    continue

                existing = self.repository.find_by_product_id(product_id)
                if existing:
                    self.vector_store.delete_product_vectors_by_product_id(product_id)
                    product = existing
                else:
                    product = Product(product_id=product_id, name=name)

                product.name = name
                product.category = self._str_or_none(row[header_map["Category"]])
                product.brand = self._str_or_none(row[header_map["Brand"]])
                product.description = self._str_or_none(row[header_map["Description"]])
                product.price = self._decimal_or_default(row[header_map["Price"]], Decimal("0.00"))
                product.image_url = self._str_or_none(row[header_map["ImageUrl"]])
                product.rating = self._decimal_or_none(row[header_map["Rating"]])
                product.stock_count = self._int_or_default(row[header_map["StockCount"]], 0)

                self.repository.create(product)

                embed_text = (
                    f"Product: {product.name}. Category: {product.category or ''}. "
                    f"Brand: {product.brand or ''}. Description: {product.description or ''}. "
                    f"Price: ${product.price}."
                )
                embedding = self.embeddings.embed(embed_text)
                self.vector_store.add_product_vector(
                    content=embed_text,
                    metadata={"product_id": product.product_id, "category": product.category or ""},
                    embedding=embedding,
                )

                imported += 1
            except Exception as exc:
                errors.append(f"Row {row_number}: {exc}")
                skipped += 1

        return ProductUploadResult(imported=imported, skipped=skipped, errors=errors)

    def search_products(self, query: str, top_k: int, threshold: float) -> list[Product]:
        """Return relational products ranked by vector similarity."""
        embedding = self.embeddings.embed(query)
        max_distance = self.vector_store.max_distance_from_similarity_threshold(threshold)
        hits = self.vector_store.search_product_vectors(
            embedding=embedding,
            top_k=top_k,
            max_distance=max_distance,
        )

        products: list[Product] = []
        for hit in hits:
            product_id = str(hit.metadata.get("product_id", "")).strip()
            if not product_id:
                continue
            product = self.repository.find_by_product_id(product_id)
            if product:
                products.append(product)

        return products

    @staticmethod
    def _str_or_none(value) -> str | None:
        """Normalize an optional spreadsheet cell to stripped text."""
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    @staticmethod
    def _decimal_or_default(value, default: Decimal) -> Decimal:
        """Parse a spreadsheet value as Decimal or return the supplied default."""
        try:
            return Decimal(str(value))
        except Exception:
            return default

    @staticmethod
    def _decimal_or_none(value):
        """Parse a nullable spreadsheet value as Decimal."""
        if value is None or str(value).strip() == "":
            return None
        try:
            return Decimal(str(value))
        except Exception:
            return None

    @staticmethod
    def _int_or_default(value, default: int) -> int:
        """Parse a spreadsheet value as int or return the supplied default."""
        try:
            return int(value)
        except Exception:
            return default
